#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
web_panel/routes/export.py - Endpoints para exportar PDF y Excel
"""

from flask import Blueprint, send_file, request, flash, redirect, url_for
from web_panel.utils.decorators import role_required
from web_panel.models.access_log import AccessLog
from web_panel.models.user import User
from config.settings import config
from datetime import datetime
import os
import io

export_bp = Blueprint('export', __name__, url_prefix='/export')

@export_bp.route('/pdf')
@role_required('admin', 'guardia')
def export_pdf():
    try:
        from fpdf import FPDF
    except ImportError:
        flash("La librería FPDF no está instalada. Ejecuta: pip install fpdf", "error")
        return redirect(url_for('dashboard.index'))
    
    # Crear reporte PDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=15, style='B')
    
    # Header
    pdf.cell(200, 10, txt=f"Reporte de Accesos - {config.SYSTEM_NAME}", ln=True, align='C')
    pdf.cell(200, 10, txt=f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ln=True, align='C')
    
    # Obtener ultimos 500 logs para reporte limitandolo por rendimiento
    logs = AccessLog.get_recent(500)
    
    pdf.set_font("Arial", size=10, style='B')
    pdf.ln(10)
    pdf.cell(50, 10, "Fecha / Hora", border=1)
    pdf.cell(100, 10, "Detalle (Usuario/QR)", border=1)
    pdf.cell(40, 10, "Resultado", border=1, ln=True)
    
    pdf.set_font("Arial", size=9)
    for log in logs:
        detalle = log.qr_texto
        if log.is_permitido():
            user = User.get_by_qr(log.qr_texto)
            if user:
                detalle = f"{user.nombre} ({user.empresa})"
        
        pdf.cell(50, 10, log.fecha_hora.strftime("%Y-%m-%d %H:%M"), border=1)
        # Cortar a 45 caracteres el detalle si es muy largo
        pdf.cell(100, 10, detalle[:45] + ('...' if len(detalle) > 45 else ''), border=1)
        pdf.cell(40, 10, log.resultado, border=1, ln=True)

    # Generar temporal 
    filename = f"Reporte_Accesos_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    filepath = config.BASE_DIR / 'logs' / filename
    
    # Ensure logs dir exists
    os.makedirs(config.BASE_DIR / 'logs', exist_ok=True)
    
    pdf.output(str(filepath))
    return send_file(str(filepath), as_attachment=True, download_name=filename)

@export_bp.route('/excel')
@role_required('admin', 'guardia')
def export_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        flash("La librería openpyxl no está instalada. Ejecuta: pip install openpyxl", "error")
        return redirect(url_for('dashboard.index'))
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Accesos QR"
    
    # Headers
    headers = ["ID", "Fecha/Hora", "Código Identificador", "Usuario Identificado", "Resultado"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Fetch data
    logs = AccessLog.get_recent(1000)
    
    row_num = 2
    for log in logs:
        ws.cell(row=row_num, column=1, value=log.id)
        ws.cell(row=row_num, column=2, value=log.fecha_hora.strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=row_num, column=3, value=log.qr_texto)
        
        usuario_nom = "Desconocido/Visitante"
        if log.is_permitido():
            user = User.get_by_qr(log.qr_texto)
            if user:
                usuario_nom = f"{user.nombre} ({user.empresa})"
        
        ws.cell(row=row_num, column=4, value=usuario_nom)
        ws.cell(row=row_num, column=5, value=log.resultado)
        row_num += 1
    
    # Autoajuste de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save to memory and send
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    filename = f"Reporte_Accesos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(excel_file, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
